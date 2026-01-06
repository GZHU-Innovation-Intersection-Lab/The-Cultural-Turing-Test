import httpx
from openai import OpenAI
from time import sleep
import openai
import pandas as pd
from openpyxl import load_workbook
from openai import OpenAI
import os
import re
from openpyxl import load_workbook, Workbook
from enum import Enum

class TestType(Enum):
    TEST_1 = "1"
    TEST_2 = "2"
    TEST_3 = "3"
    TEST_4_1 = "4.1"
    TEST_4_2 = "4.2"
    TEST_5_1 = "5.1"
    TEST_5_2 = "5.2"
    TEST_6_1 = "6.1"
    TEST_6_2 = "6.2"
    TEST_7_1 = "7.1"
    TEST_7_2 = "7.2"
    TEST_8 = "8"
    TEST_9 = "9"

    def getValue(self):
        return self.value

def get_file_path(test,country):
    file_paths = []
    if(test == "1" or test == "2"):
        file_paths = [
            fr"prompt\test{test}\CN-test1.txt",
            fr"prompt\test{test}\{country}_persona_{test}.xlsx",
            r"prompt\prompt3.xlsx"
        ]


    if (test == "3"):
        file_paths = [
            fr"prompt\test3\{country}-prompt3.1.txt",
            fr"prompt\test3\{country}-prompt3.2.txt",
            r"prompt\prompt3.xlsx"
        ]

    if (test == "4.1" or test == "4.2"):
        file_paths = [
            fr"prompt\test4\CN-test1.txt",
            fr"prompt\test4\{country}_persona_{test}.xlsx",
            r"prompt\prompt3.xlsx"
        ]

    if (test == "5.1" or test == "5.2"):
        file_paths = [
            fr"prompt\test5\{country}-test5.txt",
            fr"prompt\test5\{country}_persona_{test}.xlsx",
            r"prompt\prompt3.xlsx"
        ]

    if (test == "6.1" or test == "6.2"):
        file_paths = [
            fr"prompt\test6\CN-test1(cantonese).txt",
            fr"prompt\test6\{country}_persona_{test}.xlsx",
            r"prompt\prompt3(cantonese).xlsx"
        ]

    if (test == "7.1" or test == "7.2"):
        file_paths = [
            fr"prompt\test7\CN-test1(14-16).txt",
            fr"prompt\test7\{country}_persona_{test}.xlsx",
            r"prompt\prompt3.xlsx"
        ]

    if (test == "8" or test == "9"):
        file_paths = [
            fr"prompt\test{test}\prompt({country}).txt",
            fr"prompt\test{test}\null.txt",
            r"prompt\prompt3.xlsx"
        ]

    if(len(file_paths) == 0):
        raise ValueError(f"Invalid test value: {test}")

    return file_paths
def modelTry(test,run,country,model,base_path):
    global df
    proxy_url = "socks5://127.0.0.1:10808"
    transport = httpx.HTTPTransport(proxy=proxy_url)
    # TODO update API here
    client = OpenAI(
        api_key='sk-ioEf771d7fa104b380fc9f9aeff8a21438513a71e377OIRQ',
        base_url="https://api.deepseek.com"
    )
    file_paths = get_file_path(test,country)
    df_columns = [f"Q{i + 1}" for i in range(24)]
    df = pd.DataFrame(columns=df_columns)
    resultList = []

    def getAnswer(idx, responses):
        answers = re.findall(r"Q\d.*?[:：]\s*(\d+)", responses[idx - 3] + responses[idx - 2] + responses[idx - 1])
        if (answers == []):
            answers = re.findall(r"A\d.*?[:：]\s*(\d+)", responses[idx - 3] + responses[idx - 2] + responses[idx - 1])
            if (answers == []):
                answers = re.findall(r"Answer+.*?[:：]\s*(\d+)",
                                     responses[idx - 3] + responses[idx - 2] + responses[idx - 1])
                if (answers == []):
                    answers = re.findall(r'Q\d+[:：](\d+)', responses[idx - 3] + responses[idx - 2] + responses[idx - 1])
                    if (answers == []):
                        answers = re.findall(r'\*\*Q\d+\. .*?\n(\d+)\. ',
                                             responses[idx - 3] + responses[idx - 2] + responses[idx - 1])
                        if (answers == []):
                            answers = re.findall(r'\*\*(\d+)\.',
                                                 responses[idx - 3] + responses[idx - 2] + responses[idx - 1])
                            if (answers == []):
                                answers = re.findall(r"Choice[:：]\s*(\d+)\.",
                                                     responses[idx - 3] + responses[idx - 2] + responses[idx - 1])
                                if (answers == []):
                                    answers = re.findall(r'\*\*Q\d+\.\s+.*?\n(\d+)',
                                                         responses[idx - 3] + responses[idx - 2] + responses[idx - 1])
                                    if (answers == []):
                                        answers = re.findall(r"→ (\d+)\.",
                                                             responses[idx - 3] + responses[idx - 2] + responses[
                                                                 idx - 1])
                                    if (answers == []):
                                        answers = re.findall(r'\*\*Q\d+[:：](\d+)',
                                                                 responses[idx - 3] + responses[idx - 2] + responses[
                                                                     idx - 1])
                                        if (answers == []):
                                            answers = re.findall(r'\*\*Q\d+\*\*[:：](\d+)',
                                                                     responses[idx - 3] + responses[idx - 2] +
                                                                     responses[
                                                                         idx - 1])
        return answers

    def save_res(resultList, file_path, start):
        i = start - 1

        file_path = os.path.abspath(file_path)

        for result in resultList:
            i += 1
            answers = result.split()

            if len(answers) != 24:
                print(f"Item {i} length is incorrect, inserting blank row")
                answers = [''] * 24

            wb = None
            try:
                if os.path.exists(file_path):
                    wb = load_workbook(file_path)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Sheet"

                new_row = ws.max_row + 1

                for col_idx, answer in enumerate(answers, start=1):
                    ws.cell(row=new_row, column=col_idx, value=answer)

                wb.save(file_path)
                print(f"Row {i} saved successfully")

            except Exception as e:
                print(f"Error while saving file: {str(e)}")
                try:
                    wb = Workbook()
                    ws = wb.active
                    for col_idx, answer in enumerate(answers, start=1):
                        ws.cell(row=1, column=col_idx, value=answer)
                    wb.save(file_path)
                    print(f"Created new file and saved row {i}")
                except Exception as fallback_e:
                    print(f"Unable to create new file: {str(fallback_e)}")
            finally:
                if wb is not None:
                    wb.close()

    def read_xlsx(xlsx_path, minRow, maxCol, questions):
        wb = load_workbook(filename=xlsx_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=minRow, max_col=maxCol, max_row=ws.max_row, values_only=True):
            question = row[0] if row[0] is not None else ""
            for i in range(1, maxCol):
                if row[i] is not None:
                    question += str(i) + '.' + str(row[i])
            if question:
                questions.append(question)

    def call_deepseek_api(messages):
        try:
            response = client.chat.completions.create(
                # TODO update model name if needed
                # model="deepseek-ai/DeepSeek-R1",
                # model = "deepseek-chat",
                model="gpt-5-chat",
                messages=messages,
                # temperature=1.3,
                max_tokens=6144,
                stream=False

            )
            return response.choices[0].message.content.strip()
        except Exception as e:
            print(f"API request failed: {e}")
            return "Error"

    def estimate_token_count(text):
        return len(text.split())

    def trim_messages(messages, max_tokens=6144):
        total_tokens = sum(estimate_token_count(m['content']) for m in messages)

        while total_tokens > max_tokens and len(messages) > 1:
            removed = messages.pop(0)
            total_tokens -= estimate_token_count(removed['content'])

        return messages

    def process_prompts(window_id):
        responses = [""] * len(df_columns)
        global df
        messages_history = []
        for i, file_path in enumerate(file_paths[:2]):
            if file_path.lower().endswith('.xlsx'):
                wb = load_workbook(filename=file_path)
                ws = wb.active
                row_number = min(window_id, ws.max_row)
                cell_value = ws.cell(row=row_number, column=1).value
                prompt = str(cell_value).strip() if cell_value is not None else ""
            else:
                with open(file_path, 'r', encoding='utf-8') as file:
                    prompt = file.read().strip()
            role = "system" if i < 2 else "user"
            message = {"role": role, "content": prompt}
            messages_history.append(message)

        batch_sizes = [24]
        start_idx = 0
        idx = 0
        for batch_size in batch_sizes:
            question = questions[start_idx:start_idx + batch_size]
            start_idx += batch_size
            question = " ".join(question)
            messages_history.append({"role": "user", "content": question})
            print(f"Processing question {idx + 1}: {question}")
            trimmed_messages = messages_history.copy()
            try:
                trim = trim_messages(trimmed_messages)
                response = call_deepseek_api(trim)
                responses[idx] = response
                print(f"Question {idx + 1}: {response}")
            except Exception as e:
                print(f"Error processing question {idx + 1}: {e}")
                try:
                    print("Retrying")
                    response = call_deepseek_api(trim)
                    responses[idx] = response
                    print(f"Question {idx + 1} retry: {response}")
                except Exception as e:
                    responses[idx] = "Error"

            messages_history.append({"role": "assistant", "content": responses[idx]})
            idx += 1

        answers = getAnswer(idx, responses)
        result = " ".join(answers)
        print(result)
        resultList.append(result)
        new_row = pd.DataFrame([responses], columns=df_columns)
        df = pd.concat([df, new_row], ignore_index=True)

    questions = []
    xlsx_path = file_paths[2]
    read_xlsx(xlsx_path, 2, 6, questions)

    for window_id in range(1,21):
        process_prompts(window_id)
        print(f"Completed iteration {window_id}")
        sleep(2)

    if(len(test)>1):
        base_path = base_path + "\\test" + test[0]

    file_path = fr"{base_path}\\test{test}\{country}-test{test}({model} {run}).xlsx"

    if os.path.exists(file_path):
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            sheet_name = 'Sheet1'
            if sheet_name in writer.sheets:
                startrow = writer.sheets[sheet_name].max_row
            else:
                startrow = 0
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)
    else:
        df.to_excel(file_path, index=False)
    print(resultList)

    save_res(resultList, fr"{base_path}\\test{test}\{country}-res{test}({model} {run}).xlsx", 1)
    print("Processing complete, results saved")

base_path = "res"

for run in range(1, 6):
    # Parameters: test, run index, country (CN/HK), model name, base path
    test = TestType.TEST_1

    modelTry(test.getValue(),run,"HK","gpt-5-chat",base_path)
