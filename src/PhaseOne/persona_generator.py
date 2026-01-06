from time import sleep

import openai
import pandas as pd
from openpyxl import load_workbook
from openai import OpenAI
import os
import re


API_KEY = ''
client = OpenAI(api_key=API_KEY, base_url="https://api.gptsapi.net/v1")

file_paths = [
    fr"14-16\HK-prompt.txt",
    r"HK_character_roles.xlsx",
    r"prompt3.xlsx"
]

df_columns = [f"Q{i + 1}" for i in range(24)]
df = pd.DataFrame(columns=df_columns)
resultList = []


def getAnswer(idx, responses):
    answers = re.findall(r"Q\d.*?[:：]\s*(\d+)", responses[idx - 3] + responses[idx - 2] + responses[idx - 1])
    if (answers == []):
        answers = re.findall(r"A\d.*?[:：]\s*(\d+)", responses[idx - 3] + responses[idx - 2] + responses[idx - 1])
        if (answers == []):
            answers = re.findall(r"Answer+.*?[:：]\s*(\d+)", responses[idx - 3] + responses[idx - 2] + responses[idx - 1])
            if (answers == []):
                answers = re.findall(r'Q\d+[:：](\d+)', responses[idx - 3] + responses[idx - 2] + responses[idx - 1])
                if(answers == []):
                    answers = re.findall(r'\*\*Q\d+\. .*?\n(\d+)\. ', responses[idx - 3] + responses[idx - 2] + responses[idx - 1])
                    if (answers == []):
                        answers = re.findall(r'\*\*(\d+)\.',
                                             responses[idx - 3] + responses[idx - 2] + responses[idx - 1])
                        if (answers == []):
                            answers = re.findall(r"Choice[:：]\s*(\d+)\.",
                                                 responses[idx - 3] + responses[idx - 2] + responses[idx - 1])
                            if (answers == []):
                                answers = re.findall(r'\*\*Q\d+\.\s+.*?\n(\d+)',
                                                     responses[idx - 3] + responses[idx - 2] + responses[idx - 1])
                                if (answers == []):
                                    answers = re.findall(r"→ (\d+)\.",
                                                         responses[idx - 3] + responses[idx - 2] + responses[idx - 1])
                                    if (answers == []):
                                        answers = re.findall(r'\*\*Q\d+[:：](\d+)',
                                                             responses[idx - 3] + responses[idx - 2] + responses[
                                                                 idx - 1])
                                        if (answers == []):
                                            answers = re.findall(r'\*\*Q\d+\*\*[:：](\d+)',
                                                                 responses[idx - 3] + responses[idx - 2] + responses[
                                                                     idx - 1])
    return answers


def save_res(resultList, file_path,start):
    i=start-1
    for result in resultList:
        i+=1
        answers = result.split()
        if len(answers) != 24:
            print(f"Item {i} save failed")
            continue

        try:
            wb = load_workbook(file_path)
            ws = wb.active

            new_row = ws.max_row + 1

            for col_idx, answer in enumerate(answers, start=1):
                ws.cell(row=new_row, column=col_idx, value=answer)

            wb.save(file_path)
            print(f"Saved {answers} successfully")

        except Exception as e:
            print(f"Error while saving data: {e}")


def read_xlsx(xlsx_path,minRow,maxCol,questions):
    wb = load_workbook(filename=xlsx_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=minRow, max_col=maxCol, max_row=ws.max_row, values_only=True):
        question = row[0]
        for i in range(1, maxCol):
            question += str(i) + '.' + row[i]
        if question is not None:
            questions.append(question)


def call_deepseek_api(messages):
    try:
        response = client.chat.completions.create(
            model = "o3-mini",
            messages=messages,
            temperature=1.3,
            max_tokens=8192,
            stream=False
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        print(f"API request failed: {e}")
        return "Error"


def estimate_token_count(text):
    return len(text.split())


def trim_messages(messages, max_tokens=6144):
    total_tokens = sum(estimate_token_count(m['content']) for m in messages)

    while total_tokens > max_tokens and len(messages) > 1:
        removed = messages.pop(0)
        total_tokens -= estimate_token_count(removed['content'])

    return messages


def process_prompts(window_id):
    global df
    responses = [""] * len(df_columns)

    messages_history = []
    for i, file_path in enumerate(file_paths[:2]):
        if file_path.lower().endswith('.xlsx'):
            wb = load_workbook(filename=file_path)
            ws = wb.active
            row_number = min(window_id, ws.max_row)
            cell_value = ws.cell(row=row_number, column=1).value
            prompt = str(cell_value).strip() if cell_value is not None else ""
        else:
            with open(file_path, 'r', encoding='utf-8') as file:
                prompt = file.read().strip()
        role = "system" if i <2 else "user"
        if(i == 1 ):
            message = {"role": role, "content": "The role to generate is: " + prompt}
        else:
            message = {"role": role, "content": prompt}
        messages_history.append(message)

    batch_sizes = [1]
    start_idx = 0
    idx=0
    for batch_size in batch_sizes:
        question = questions[start_idx:start_idx + batch_size]
        start_idx += batch_size
        question = " ".join(question)
        messages_history.append({"role": "user", "content": question})
        print(f"Processing question {idx+1}: {question}")
        trimmed_messages = messages_history.copy()
        try:
            trim = trim_messages(trimmed_messages)
            response = call_deepseek_api(trim)
            responses[idx] = response
            print(f"Question {idx + 1}: {response}")
        except Exception as e:
            print(f"Error processing question {idx + 1}: {e}")
            try:
                print("Retrying")
                response = call_deepseek_api(trim)
                responses[idx] = response
                print(f"Question {idx + 1} retry: {response}")
            except Exception as e:
                responses[idx] = "Error"

        messages_history.append({"role": "assistant", "content": responses[idx]})
        idx+=1

    answers = getAnswer(idx, responses)
    result = " ".join(answers)
    print(result)
    resultList.append(result)
    new_row = pd.DataFrame([responses], columns=df_columns)
    df = pd.concat([df, new_row], ignore_index=True)


pro = "Please generate a persona"
questions = []
questions.append(pro)

xlsx_path = file_paths[2]
read_xlsx(xlsx_path,2,6,questions)

for window_id in range(1,41):
    process_prompts(window_id)
    print(f"Completed iteration {window_id}")
    sleep(2)

file_path = r"14-16\HK_persona_1.xlsx"

if os.path.exists(file_path):
    with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
        sheet_name = 'Sheet1'
        if sheet_name in writer.sheets:
            startrow = writer.sheets[sheet_name].max_row
        else:
            startrow = 0
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)
else:
    df.to_excel(file_path, index=False)

print("Processing complete, results saved")

