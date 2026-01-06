import os
import re
import json
import httpx
import pandas as pd
from time import sleep
from datetime import datetime, timezone
from openpyxl import load_workbook, Workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
from openai import OpenAI


def get_file_config():
    return {
        "prompt_file": r"prompt\test2\test1.txt",  # Prompt only
        "question_file": r"prompt\prompt3.xlsx",      # Questions only
    }


def ensure_dir(path):
    if not os.path.exists(path):
        os.makedirs(path, exist_ok=True)


def read_xlsx_questions(xlsx_path, min_row=2, max_col=6, limit_count=24):
    questions = []
    wb = load_workbook(filename=xlsx_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=min_row, max_col=max_col, max_row=ws.max_row, values_only=True):
        question = row[0] if row[0] is not None else ""
        for i in range(1, max_col):
            if row[i] is not None:
                question += str(i) + '.' + str(row[i])
        if question and str(question).strip():
            questions.append(str(question))
        if limit_count and len(questions) >= limit_count:
            break
    return questions


def load_text(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        return f.read().strip()


def create_client(api_key, base_url, proxy_enabled=False, proxy_url='http://127.0.0.1:7890'):
    http_client = None
    if proxy_enabled:
        transport = httpx.HTTPTransport(proxy=proxy_url)
        http_client = httpx.Client(transport=transport)

    client = OpenAI(
        api_key=api_key,
        base_url=base_url,
        http_client=http_client
    )
    return client


def call_model(client, model, messages, max_tokens=6144, timeout=120, retries=3):
    last_err = None
    for attempt in range(1, retries + 1):
        try:
            resp = client.chat.completions.create(
                model=model,
                messages=messages,
                max_tokens=max_tokens,
                stream=False,
                timeout=timeout
            )
            content = resp.choices[0].message.content.strip()
            return {"ok": True, "content": content, "error": ""}
        except Exception as e:
            last_err = str(e)
            print(f"{model} Request failed (Attempt {attempt}): {e}")
            sleep(min(5 * attempt, 15))
    return {"ok": False, "content": "", "error": last_err or "Unknown error"}


def extract_numeric_answers(blocks, max_count=24):
    text = "\n".join(blocks)
    patterns = [
        r"Q\d.*?[:：]\s*(\d+)",
        r"A\d.*?[:：]\s*(\d+)",
        r"Answer+.*?[:：]\s*(\d+)",
        r"Q\d+：(\d+)",
        r"**Q\d+\. .*?\n(\d+)\. ",
        r"**(\d+)\.",
        r"Selection: (\d+)\.",
        r"\*\*Q\d+\.\s+.*?\n(\d+)",
        r"→ (\d+)\.",
        r"\*\*Q\d+：(\d+)",
        r"\*\*Q\d+\*\*：(\d+)",
    ]
    for pat in patterns:
        answers = re.findall(pat, text)
        if answers:
            return answers[:max_count]
    return []


def save_numeric_answers(numbers_list, file_path, start_index=1, max_count=24):
    ensure_dir(os.path.dirname(file_path))
    file_path = os.path.abspath(file_path)
    wb = None
    try:
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet"
        row_cursor = ws.max_row + 1 if ws.max_row > 1 or ws.cell(1, 1).value else 1
        i = start_index - 1
        for numbers in numbers_list:
            i += 1
            answers = list(numbers)
            if len(answers) < max_count:
                answers = answers + [''] * (max_count - len(answers))
            else:
                answers = answers[:max_count]
            for col_idx, val in enumerate(answers, start=1):
                ws.cell(row=row_cursor, column=col_idx, value=val)
            row_cursor += 1
        wb.save(file_path)
    finally:
        if wb is not None:
            wb.close()


def save_persona_all_answers(persona_row, answers, file_path):
    """Save all answers for a single persona, one persona per row."""
    ensure_dir(os.path.dirname(file_path))
    file_path = os.path.abspath(file_path)
    wb = None
    try:
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet"
            # Write header
            ws.cell(row=1, column=1, value="PersonaRow")
            for i in range(1, 25):  # Q1 to Q24
                ws.cell(row=1, column=i+1, value=f"Q{i}")
        
        row_cursor = ws.max_row + 1 if ws.max_row > 1 or ws.cell(1, 1).value else 2
        ws.cell(row=row_cursor, column=1, value=persona_row)
        
        # Fill all answers
        for i, answer in enumerate(answers):
            if i < 24:  # Max 24 answers
                ws.cell(row=row_cursor, column=i+2, value=answer)
        
        wb.save(file_path)
    finally:
        if wb is not None:
            wb.close()


def append_jsonl(file_path, obj):
    ensure_dir(os.path.dirname(file_path))
    with open(file_path, 'a', encoding='utf-8') as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")


def append_csv_row(file_path, headers, values):
    import csv
    ensure_dir(os.path.dirname(file_path))
    file_exists = os.path.exists(file_path)
    with open(file_path, 'a', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(headers)
        writer.writerow(values)


def append_persona_all_answers_csv(file_path, persona_row, answers):
    """Save all answers for a single persona to CSV, one persona per row."""
    import csv
    ensure_dir(os.path.dirname(file_path))
    file_exists = os.path.exists(file_path)
    with open(file_path, 'a', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        if not file_exists:
            headers = ["PersonaRow"] + [f"Q{i}" for i in range(1, 25)]
            writer.writerow(headers)
        
        # Prepare data row: persona_row + 24 answers
        row_data = [persona_row] + (answers + [''] * (24 - len(answers)))[:24]
        writer.writerow(row_data)


SAVE_SEPARATE = True  # Save separately, do not append to old file


def parse_personas_from_text(text, expect_count=42):
    # Try parsing JSON array
    try:
        data = json.loads(text)
        if isinstance(data, list):
            personas = [str(x).strip() for x in data if str(x).strip()]
            if personas:
                return personas[:expect_count]
    except Exception:
        pass
    # Fallback: Parse by line/number
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    personas = []
    for line in lines:
        # Matches "1. XXX" or "- XXX"
        m = re.match(r"^(?:\d+\.|[-•])\s*(.+)$", line)
        if m:
            personas.append(m.group(1))
        else:
            personas.append(line)
        if len(personas) >= expect_count:
            break
    return personas[:expect_count]


def generate_personas(client, model_name, base_prompt_text):
    # Use your prompt directly, no extra instructions
    messages = [
        {"role": "user", "content": base_prompt_text},
    ]
    result = call_model(client, model_name, messages)
    content = result["content"] if result.get("ok") else ""
    personas = parse_personas_from_text(content, expect_count=42)
    status = "success" if personas else "error"
    return {
        "ok": bool(personas),
        "status": status,
        "error": result.get("error", "") if not personas else "",
        "raw": content,
        "personas": personas,
    }


def run_pipeline_for_model(model_name, api_key, base_url, out_dir):
    files = get_file_config()
    prompt_text = load_text(files["prompt_file"])  # Prompt only
    questions = read_xlsx_questions(files["question_file"], min_row=2, max_col=6, limit_count=24)
    questions_join = " ".join(questions)

    client = create_client(api_key=api_key, base_url=base_url, proxy_enabled=False)

    ensure_dir(out_dir)
    ts_suffix = ("." + datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S') + ".gen") if SAVE_SEPARATE else ""

    # Output files (this batch)
    personas_xlsx = os.path.join(out_dir, f"personas({model_name}){ts_suffix}.xlsx")
    personas_jsonl = os.path.join(out_dir, f"personas({model_name}){ts_suffix}.jsonl")
    eval_text_xlsx = os.path.join(out_dir, f"eval({model_name}){ts_suffix}.xlsx")
    eval_text_jsonl = os.path.join(out_dir, f"eval({model_name}){ts_suffix}.jsonl")
    eval_num_xlsx = os.path.join(out_dir, f"res({model_name}){ts_suffix}.xlsx")
    eval_num_csv = os.path.join(out_dir, f"res({model_name}){ts_suffix}.csv")
    # Save all answers by persona
    persona_answers_xlsx = os.path.join(out_dir, f"persona-answers({model_name}){ts_suffix}.xlsx")
    persona_answers_csv = os.path.join(out_dir, f"persona-answers({model_name}){ts_suffix}.csv")

    # 1) Generate and save personas (using your prompt directly)
    gen = generate_personas(client, model_name, prompt_text)
    if gen.get("ok"):
        # Save to xlsx
        wb = Workbook()
        ws = wb.active
        ws.title = "Personas"
        ws.cell(row=1, column=1, value="Persona")
        r = 2
        for p in gen["personas"]:
            ws.cell(row=r, column=1, value=p)
            r += 1
        wb.save(personas_xlsx)
        wb.close()
        # Save to jsonl
        for idx, p in enumerate(gen["personas"], 1):
            append_jsonl(personas_jsonl, {
                "ts": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
                "model": model_name,
                "index": idx,
                "persona": p,
            })
    else:
        # Write error record if failed
        append_jsonl(personas_jsonl, {
            "ts": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
            "model": model_name,
            "status": "error",
            "error": gen.get("error", "generate personas failed")
        })
        print(f"{model_name} Failed to generate personas: {gen.get('error','')}")
        return {
            "model": model_name,
            "personas_file": personas_xlsx,
            "text_file": eval_text_xlsx,
            "numeric_file": eval_num_xlsx
        }

    # 2) Answer per persona and save immediately (save generated personas even if interrupted)
    for row_idx, persona_text in enumerate(gen["personas"], 1):
        messages = [
            {"role": "system", "content": prompt_text},
            {"role": "system", "content": persona_text},
            {"role": "user", "content": questions_join},
        ]
        result = call_model(client, model_name, messages)
        response_text = result["content"] if result.get("ok") else f"Error: {result.get('error','')}"

        # Save text answer (Excel incremental)
        try:
            df_one = pd.DataFrame([{ "PersonaRow": row_idx, "Status": "success" if result.get("ok") else "error", "Response": response_text }])
            if os.path.exists(eval_text_xlsx):
                with pd.ExcelWriter(eval_text_xlsx, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                    sheet_name = 'Sheet1'
                    startrow = writer.sheets[sheet_name].max_row if sheet_name in writer.sheets else 0
                    df_one.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)
            else:
                df_one.to_excel(eval_text_xlsx, index=False)
        except Exception as e:
            print(f"Failed to save text answer (Persona {row_idx}): {e}")

        # Save text answer (JSONL incremental)
        try:
            append_jsonl(eval_text_jsonl, {
                "ts": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
                "model": model_name,
                "PersonaRow": row_idx,
                "Status": "success" if result.get("ok") else "error",
                "Response": response_text
            })
        except Exception as e:
            print(f"Failed to save JSONL (Persona {row_idx}): {e}")

        # Save numeric answers (Excel/CSV incremental)
        try:
            numbers = extract_numeric_answers([response_text], max_count=24) if result.get("ok") else []
            save_numeric_answers([numbers], eval_num_xlsx, start_index=1, max_count=24)
            headers = ["PersonaRow"] + [f"Q{i}" for i in range(1, 25)]
            row_vals = [row_idx] + (numbers + [''] * (24 - len(numbers)))[:24]
            append_csv_row(eval_num_csv, headers, row_vals)
            
            # Save all answers by persona
            save_persona_all_answers(row_idx, numbers, persona_answers_xlsx)
            append_persona_all_answers_csv(persona_answers_csv, row_idx, numbers)
        except Exception as e:
            print(f"Failed to save numeric answers (Persona {row_idx}): {e}")

        print(f"{model_name} Persona {row_idx} Completed")
        sleep(2)

    return {
        "model": model_name,
        "personas_file": personas_xlsx,
        "text_file": eval_text_xlsx,
        "numeric_file": eval_num_xlsx,
        "persona_answers_file": persona_answers_xlsx
    }


def main():
    # Model and API configuration
    default_api_key = os.getenv("OPENAI_API_KEY", '')
    default_base_url = os.getenv("OPENAI_BASE_URL", "https://api.gptsapi.net/v1")

    models = [
        "gpt-5-chat",
    ]

    model_overrides = {}

    out_dir = os.path.join("res", "multi-gen")
    ensure_dir(out_dir)

    results = []
    with ThreadPoolExecutor(max_workers=len(models)) as executor:
        futures = []
        for m in models:
            cfg = model_overrides.get(m, None)
            api_key = cfg["api_key"] if cfg else default_api_key
            base_url = cfg["base_url"] if cfg else default_base_url
            futures.append(executor.submit(run_pipeline_for_model, m, api_key, base_url, out_dir))
        for f in as_completed(futures):
            try:
                results.append(f.result())
            except Exception as e:
                print(f"Thread Exception: {e}")

    for r in results:
        print(f"{r['model']} => Personas: {r['personas_file']} | Text: {r['text_file']} | Numeric: {r['numeric_file']} | Persona Answers: {r['persona_answers_file']}")


if __name__ == "__main__":
    main()
