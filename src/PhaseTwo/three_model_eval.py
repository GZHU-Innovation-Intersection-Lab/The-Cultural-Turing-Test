import os
import re
import json
import httpx
import pandas as pd
from time import sleep
from datetime import datetime, timezone
from openpyxl import load_workbook, Workbook
from openai import OpenAI


# Unified file configuration: three threads share the same prompt, persona, and questions
def get_file_config():
    return {
        "prompt_file": r"prompt\test1\test1.txt",
        "character_file": r"prompt\test1\persona1.xlsx",
        "question_file": r"prompt\prompt3.xlsx",
    }


def ensure_dir(path):
    if not os.path.exists(path):
        os.makedirs(path, exist_ok=True)


# Whether to "save as" new file for this run (do not append to old file)
SAVE_SEPARATE = True


def read_xlsx_questions(xlsx_path, min_row, max_col, limit_count=None):
    questions = []
    wb = load_workbook(filename=xlsx_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=min_row, max_col=max_col, max_row=ws.max_row, values_only=True):
        question = row[0] if row[0] is not None else ""
        for i in range(1, max_col):
            if row[i] is not None:
                question += str(i) + '.' + str(row[i])
        if question and str(question).strip():
            questions.append(str(question))
        if limit_count and len(questions) >= limit_count:
            break
    return questions


def load_text(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        return f.read().strip()


def load_persona_row(xlsx_path, row_number):
    wb = load_workbook(filename=xlsx_path)
    ws = wb.active
    row_number = min(max(1, row_number), ws.max_row)
    cell_value = ws.cell(row=row_number, column=1).value
    return str(cell_value).strip() if cell_value is not None else ""


def extract_numeric_answers(blocks, max_count=24):
    """Extract option numbers from model returned text (robustly), returning at most max_count."""
    text = "\n".join(blocks)
    patterns = [
        r"Q\d.*?[:：]\s*(\d+)",
        r"A\d.*?[:：]\s*(\d+)",
        r"Answer+.*?[:：]\s*(\d+)",
        r"Q\d+：(\d+)",
        r"**Q\d+\. .*?\n(\d+)\. ",
        r"**(\d+)\.",
        r"Selection: (\d+)\.",
        r"\*\*Q\d+\.\s+.*?\n(\d+)",
        r"→ (\d+)\.",
    ]
    for pat in patterns:
        answers = re.findall(pat, text)
        if answers:
            return answers[:max_count]
    return []


def save_numeric_answers(numbers_list, file_path, start_index=1, max_count=24):
    """Write multiple numeric answers to Excel, each with max_count columns, padding with empty strings."""
    ensure_dir(os.path.dirname(file_path))
    file_path = os.path.abspath(file_path)

    wb = None
    try:
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet"

        row_cursor = ws.max_row + 1 if ws.max_row > 1 or ws.cell(1, 1).value else 1

        i = start_index - 1
        for numbers in numbers_list:
            i += 1
            answers = list(numbers)
            if len(answers) < max_count:
                answers = answers + [''] * (max_count - len(answers))
            else:
                answers = answers[:max_count]
            for col_idx, val in enumerate(answers, start=1):
                ws.cell(row=row_cursor, column=col_idx, value=val)
            row_cursor += 1

        wb.save(file_path)
    finally:
        if wb is not None:
            wb.close()


def append_jsonl(file_path, obj):
    """Append object to JSONL file."""
    ensure_dir(os.path.dirname(file_path))
    with open(file_path, 'a', encoding='utf-8') as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")


def append_csv_row(file_path, headers, values):
    """Append a row to CSV, write header if file does not exist."""
    import csv
    ensure_dir(os.path.dirname(file_path))
    file_exists = os.path.exists(file_path)
    with open(file_path, 'a', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(headers)
        writer.writerow(values)


def get_processed_persona_rows(text_excel_path):
    """Read existing text result table, return set of completed PersonaRows."""
    processed = set()
    try:
        if not os.path.exists(text_excel_path):
            return processed
        wb = load_workbook(filename=text_excel_path)
        ws = wb.active
        # Read header, locate PersonaRow column
        headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        if not headers:
            return processed
        try:
            col_idx = headers.index("PersonaRow") + 1
        except ValueError:
            # Default to first column if header not found
            col_idx = 1
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=col_idx).value
            if val is None:
                continue
            try:
                processed.add(int(val))
            except Exception:
                continue
        wb.close()
    except Exception:
        pass
    return processed


def create_client(api_key, base_url, proxy_enabled=False, proxy_url='http://127.0.0.1:7890'):
    http_client = None
    if proxy_enabled:
        transport = httpx.HTTPTransport(proxy=proxy_url)
        http_client = httpx.Client(transport=transport)

    client = OpenAI(
        api_key=api_key,
        base_url=base_url,
        http_client=http_client
    )
    return client


def call_model(client, model, messages, max_tokens=6144, timeout=120, retries=3):
    last_err = None
    for attempt in range(1, retries + 1):
        try:
            resp = client.chat.completions.create(
                model=model,
                messages=messages,
                max_tokens=max_tokens,
                stream=False,
                timeout=timeout
            )
            content = resp.choices[0].message.content.strip()
            return {"ok": True, "content": content, "error": ""}
        except Exception as e:                        
            last_err = str(e)
            print(f"{model} Request failed (Attempt {attempt}): {e}")
            # Exponential backoff
            sleep(min(5 * attempt, 15))
    return {"ok": False, "content": "", "error": last_err or "Unknown error"}


def run_eval_for_model(model_name, api_key, base_url, out_dir):
    files = get_file_config()
    prompt_text = load_text(files["prompt_file"])
    # 24 questions (from row 2, take first 6 columns, limit to 24 questions)
    questions = read_xlsx_questions(files["question_file"], min_row=2, max_col=6, limit_count=24)
    questions_join = " ".join(questions)

    client = create_client(api_key=api_key, base_url=base_url, proxy_enabled=False)

    responses_rows = []  # One row per persona (memory copy)
    numeric_rows = []    # Numeric answer list (memory copy)

    # Prepare output files for this model (incremental write)
    ensure_dir(out_dir)
    ts_suffix = ("." + datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S') + ".extra") if SAVE_SEPARATE else ""
    text_out = os.path.join(out_dir, f"eval({model_name}){ts_suffix}.xlsx")
    num_out = os.path.join(out_dir, f"res({model_name}){ts_suffix}.xlsx")
    # Extra save: JSONL and CSV (more robust, easier to resume)
    text_jsonl_out = os.path.join(out_dir, f"eval({model_name}){ts_suffix}.jsonl")
    num_csv_out = os.path.join(out_dir, f"res({model_name}){ts_suffix}.csv")

    # Read completed rows, only run missing ones
    processed = get_processed_persona_rows(text_out)

    # 42 personas (read first 42 rows of column 1), skip completed
    for persona_row in range(1, 43):
        if persona_row in processed:
            continue
        persona_text = load_persona_row(files["character_file"], persona_row)
        messages = [
            {"role": "system", "content": prompt_text},
            {"role": "system", "content": persona_text},
            {"role": "user", "content": questions_join},
        ]

        result = call_model(client, model_name, messages)
        response_text = result["content"] if result.get("ok") else f"Error: {result.get('error','')}"
        row_obj = {
            "PersonaRow": persona_row,
            "Status": "success" if result.get("ok") else "error",
            "Response": response_text
        }
        responses_rows.append(row_obj)

        numbers = extract_numeric_answers([response_text], max_count=24) if result.get("ok") else []
        numeric_rows.append(numbers)

        print(f"{model_name} Persona {persona_row} Completed")
        sleep(2)

        # Incremental save: Text answer (Excel)
        try:
            df_one = pd.DataFrame([row_obj])
            if os.path.exists(text_out):
                with pd.ExcelWriter(text_out, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                    sheet_name = 'Sheet1'
                    startrow = writer.sheets[sheet_name].max_row if sheet_name in writer.sheets else 0
                    df_one.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)
            else:
                df_one.to_excel(text_out, index=False)
        except Exception as e:
            print(f"Failed to append text answer (Persona {persona_row}): {e}")

        # Extra save: Text answer (JSONL)
        try:
            append_jsonl(text_jsonl_out, {
                "ts": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
                "model": model_name,
                **row_obj
            })
        except Exception as e:
            print(f"Failed to append JSONL (Persona {persona_row}): {e}")

        # Incremental save: Numeric answers (Excel, append one row at a time)
        try:
            save_numeric_answers([numbers], num_out, start_index=1, max_count=24)
        except Exception as e:
            print(f"Failed to append numeric answers (Persona {persona_row}): {e}")

        # Extra save: Numeric answers (CSV, append one row at a time)
        try:
            headers = ["PersonaRow"] + [f"Q{i}" for i in range(1, 25)]
            row_vals = [persona_row] + (numbers + [''] * (24 - len(numbers)))[:24]
            append_csv_row(num_csv_out, headers, row_vals)
        except Exception as e:
            print(f"Failed to append CSV (Persona {persona_row}): {e}")

    # Already saved incrementally in loop, just return paths

    return {
        "model": model_name,
        "text_file": text_out,
        "numeric_file": num_out
    }
