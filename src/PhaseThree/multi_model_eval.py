import os
import re
import sys
import httpx
import pandas as pd
from time import sleep
from openpyxl import load_workbook, Workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
from openai import OpenAI


def choose_existing_path(candidates):
    for p in candidates:
        if os.path.exists(p):
            return p
    return candidates[0]


def get_file_config():
    return {
        "prompt_file": choose_existing_path([r"prompt\test1\CN-test1.txt"]),
        "character_file": choose_existing_path([
            r"prompt\test1\CN_persona_1.xlsx",
            "prompt\\test1\\CN\u4eba\u8bbe1.xlsx",
        ]),
        "question_file": choose_existing_path([r"prompt\prompt3.xlsx"]),
    }


def ensure_dir(path):
    if not os.path.exists(path):
        os.makedirs(path, exist_ok=True)


def read_xlsx_questions(xlsx_path, min_row, max_col, limit_count=None):
    questions = []
    wb = load_workbook(filename=xlsx_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=min_row, max_col=max_col, max_row=ws.max_row, values_only=True):
        question = row[0] if row[0] is not None else ""
        for i in range(1, max_col):
            if row[i] is not None:
                question += str(i) + '.' + str(row[i])
        if question and str(question).strip():
            questions.append(str(question))
        if limit_count and len(questions) >= limit_count:
            break
    return questions


def load_text(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        return f.read().strip()


def load_persona_row(xlsx_path, row_number):
    wb = load_workbook(filename=xlsx_path)
    ws = wb.active
    row_number = min(max(1, row_number), ws.max_row)
    cell_value = ws.cell(row=row_number, column=1).value
    return str(cell_value).strip() if cell_value is not None else ""


def extract_numeric_answers(blocks, max_count=24):
    text = "\n".join(blocks)
    patterns = [
        r"Q\d.*?[:：]\s*(\d+)",
        r"A\d.*?[:：]\s*(\d+)",
        "\u7b54\u6848+.*?[:：]\\s*(\\d+)",
        r"Q\d+：(\d+)",
        r"\*\*Q\d+\. .*?\n(\d+)\. ",
        r"\*\*(\d+)\.",
        "\u9009\u62e9[:\uff1a](\\d+)\\.",
        r"\*\*Q\d+\.\s+.*?\n(\d+)",
        r"→ (\d+)\.",
        r"\*\*Q\d+：(\d+)",
        r"\*\*Q\d+\*\*：(\d+)",
    ]
    for pat in patterns:
        answers = re.findall(pat, text)
        if answers:
            return answers[:max_count]
    return []


def save_numeric_answers(numbers_list, file_path, start_index=1, max_count=24):
    ensure_dir(os.path.dirname(file_path))
    file_path = os.path.abspath(file_path)

    wb = None
    try:
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet"

        row_cursor = ws.max_row + 1 if ws.max_row > 1 or ws.cell(1, 1).value else 1

        i = start_index - 1
        for numbers in numbers_list:
            i += 1
            answers = list(numbers)
            if len(answers) < max_count:
                answers = answers + [''] * (max_count - len(answers))
            else:
                answers = answers[:max_count]
            for col_idx, val in enumerate(answers, start=1):
                ws.cell(row=row_cursor, column=col_idx, value=val)
            row_cursor += 1

        wb.save(file_path)
    finally:
        if wb is not None:
            wb.close()


def create_client(api_key, base_url, proxy_enabled=False, proxy_url='http://127.0.0.1:7890'):
    http_client = None
    if proxy_enabled:
        transport = httpx.HTTPTransport(proxy=proxy_url)
        http_client = httpx.Client(transport=transport)

    client = OpenAI(
        api_key=api_key,
        base_url=base_url,
        http_client=http_client
    )
    return client


def call_model(client, model, messages, max_tokens=6144, timeout=60):
    try:
        resp = client.chat.completions.create(
            model=model,
            messages=messages,
            max_tokens=max_tokens,
            stream=False,
            timeout=timeout
        )
        return resp.choices[0].message.content.strip()
    except Exception as e:
        print(f"{model} request failed: {e}")
        sleep(5)
        return "Error"


def run_eval_for_model(model_name, api_key, base_url, out_dir):
    files = get_file_config()
    prompt_text = load_text(files["prompt_file"])
    questions = read_xlsx_questions(files["question_file"], min_row=2, max_col=6, limit_count=24)
    questions_join = " ".join(questions)

    client = create_client(api_key=api_key, base_url=base_url, proxy_enabled=False)

    responses_rows = []
    numeric_rows = []

    ensure_dir(out_dir)
    text_out = os.path.join(out_dir, f"CN-eval({model_name}).xlsx")
    num_out = os.path.join(out_dir, f"CN-res({model_name}).xlsx")

    for persona_row in range(1, 41):
        persona_text = load_persona_row(files["character_file"], persona_row)
        messages = [
            {"role": "system", "content": prompt_text},
            {"role": "system", "content": persona_text},
            {"role": "user", "content": questions_join},
        ]

        result = call_model(client, model_name, messages)
        row_obj = {"PersonaRow": persona_row, "Response": result}
        responses_rows.append(row_obj)

        numbers = extract_numeric_answers([result], max_count=24)
        numeric_rows.append(numbers)

        print(f"{model_name} persona {persona_row} done")
        sleep(2)

        try:
            df_one = pd.DataFrame([row_obj])
            if os.path.exists(text_out):
                with pd.ExcelWriter(text_out, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                    sheet_name = 'Sheet1'
                    startrow = writer.sheets[sheet_name].max_row if sheet_name in writer.sheets else 0
                    df_one.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)
            else:
                df_one.to_excel(text_out, index=False)
        except Exception as e:
            print(f"Append save failed for text response (persona {persona_row}): {e}")

        try:
            save_numeric_answers([numbers], num_out, start_index=1, max_count=24)
        except Exception as e:
            print(f"Append save failed for numeric answers (persona {persona_row}): {e}")

    return {
        "model": model_name,
        "text_file": text_out,
        "numeric_file": num_out
    }


def main():
    print("Program started...")
    sys.stdout.flush()

    default_api_key = os.environ.get("DEFAULT_API_KEY", "").strip()
    default_base_url = os.environ.get("DEFAULT_BASE_URL", "https://api.gptsapi.net/v1").strip()

    models = [
        "qwen3-max-preview",
        "gpt-5-chat",
        "gemini-2.5-pro",
    ]

    print(f"Models to test: {', '.join(models)}")
    sys.stdout.flush()

    model_overrides = {
        "qwen3-max-preview": {
            "api_key": os.environ.get("QWEN_API_KEY", "").strip(),
            "base_url": os.environ.get("QWEN_BASE_URL", "https://dashscope.aliyuncs.com/compatible-mode/v1").strip(),
        }
    }

    out_dir = os.path.join("res", "multi")
    ensure_dir(out_dir)
    print(f"Output directory: {out_dir}")
    sys.stdout.flush()

    print("Starting multi-thread processing...")
    sys.stdout.flush()
    results = []
    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = []
        for m in models:
            cfg = model_overrides.get(m, None)
            api_key = (cfg.get("api_key") if cfg else "") or default_api_key
            base_url = (cfg.get("base_url") if cfg else "") or default_base_url
            if not api_key:
                raise ValueError(f"Missing API key for model '{m}'. Set DEFAULT_API_KEY or a model-specific key.")
            print(f"Submitting job: {m}")
            sys.stdout.flush()
            futures.append(executor.submit(run_eval_for_model, m, api_key, base_url, out_dir))

        print("Waiting for jobs to complete...")
        sys.stdout.flush()
        for i, f in enumerate(as_completed(futures)):
            try:
                result = f.result()
                results.append(result)
                print(f"Job {i+1}/3 done: {result['model']}")
                sys.stdout.flush()
            except Exception as e:
                print(f"Thread error: {e}")
                sys.stdout.flush()

    print("\n=== All jobs completed ===")
    sys.stdout.flush()
    for r in results:
        print(f"{r['model']} => text: {r['text_file']} | numeric: {r['numeric_file']}")

    print("Program finished.")
    sys.stdout.flush()


if __name__ == "__main__":
    main()
